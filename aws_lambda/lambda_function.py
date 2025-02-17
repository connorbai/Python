import psycopg2
import openpyxl
from openpyxl.styles import Font
import os
import uuid
import boto3
import json
import re

# from model import QueryModel
# os.environ["database"] = "idoctor_compare"

s3 = boto3.resource(service_name='s3',
                    region_name=os.environ["region_name"],
                    aws_access_key_id=os.environ["s3_key"],
                    aws_secret_access_key=os.environ["s3_secret"])


def upload_single_file_to_s3(s3_key: str, file_path: str):
    print("begin to upload {0},key:{1}", file_path, s3_key)
    bucket_name = os.environ["bucket_name"]
    with open(file_path, 'rb') as f:
        file_stream = f.read()
        try:
            s3.Bucket(bucket_name).put_object(
                Key=s3_key, Body=file_stream)
            print("{} upload done".format(s3_key))
        except Exception as e:
            print("upload {} error:{}".format(s3_key, e))


class Product:
    def __init__(self, d):
        self.__dict__ = d


def json2Product(d):
    return Product(d)


def export_to_excel(querys=[], folder_path=''):
    print("export_to_excel")
    databases = json.loads(os.environ["database"], object_hook=json2Product).data
    users = json.loads(os.environ["user"], object_hook=json2Product).data
    pwds = json.loads(os.environ["password"], object_hook=json2Product).data
    hosts = json.loads(os.environ["host"], object_hook=json2Product).data
    port = os.environ["port"]
    for database in databases:
        try:

            host = [host for host in hosts if host.database == database][0].host
            pwd = [pwd for pwd in pwds if pwd.database == database][0].pwd
            user = [user for user in users if user.database == database][0].userName
            conn = psycopg2.connect(database=database, user=user,
                                    password=pwd, host=host, port=port)
            print("Opened database successfully")
            cur = conn.cursor()
            currentQueries = [query for query in querys if query["database"] == database]
            for query in currentQueries:
                try:
                    cur.execute(query["query"])
                    data = cur.fetchall()
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    # col_names = []
                    for index in range(len(cur.description)):
                        # col_names.append()
                        sheet.cell(row=1, column=index + 1,
                                   value=cur.description[index].name)
                    # column_names = [row[0] for row in cur ]
                    for rowno, row in enumerate(data, start=1):
                        for colno, cell_value in enumerate(row, start=1):
                            # 处理NULL为NULL
                            if (cell_value == None):
                                print("67")
                                sheet.cell(row=rowno + 1, column=colno, value='NULL')
                            else:  # 避免转换时候把空值转为字符串
                                if (isinstance(cell_value, str) and cell_value.lower() == "NULL".lower()):
                                    print("71")
                                    sheet.cell(row=rowno + 1, column=colno,
                                               value='NULLStr')
                                else:  # 正常存储
                                    if (type(cell_value) == bytes):
                                        cell_value = "binary"
                                    # 处理含有\ 特殊字符
                                    # 处理特殊字符并去除添加的引号
                                    if (isinstance(cell_value, str)):
                                        cell_value = repr(cell_value)[1:-1]
                                    # #bool to str
                                    # if(isinstance(cell_value,bool)):
                                    #     cell_value=str(cell_value)
                                    print("82")
                                    sheet.cell(row=rowno + 1, column=colno,
                                               value=cell_value)

                            # .value = cell_value
                    sheet.row_dimensions[1].font = Font(bold=True)
                    if query['name']:
                        sheet.title = query['name']
                    # 同名excel文件不让同时读取 即使在不同目录
                    file_path = folder_path + "/" + query['name'] + '-pg.xlsx'
                    print(file_path)
                    create_result = wb.save(file_path)
                    s3_key = "public/tmp/dbcompare/{}".format(
                        query['name'] + '-pg.xlsx')
                    upload_single_file_to_s3(
                        s3_key=s3_key, file_path=file_path)
                except Exception as e1:
                    print("table{} row:{} col:{} failed".format(
                        query["name"], rowno, colno))
                    pass
        except Exception as e:
            print(e)
        finally:
            cur.close()
            conn.close()


def export_to_db_schema(folder_path: str):
    databases = json.loads(os.environ["database"], object_hook=json2Product).data
    users = json.loads(os.environ["user"], object_hook=json2Product).data
    pwds = json.loads(os.environ["password"], object_hook=json2Product).data
    hosts = json.loads(os.environ["host"], object_hook=json2Product).data
    port = os.environ["port"]
    for database in databases:
        try:
            query = {
                "name": "Schema",
                "query": """
                select   col.table_schema||'.'||col.table_name tablename, lower(col.column_name)  colname
                    from information_schema.columns col
                    join information_schema.tables tab on col.table_name=tab.table_name
                    where
                        tab.table_type='BASE TABLE' and
                        col.table_name similar to '[a-zA-Z_]%'
                        and col.table_name not ilike '%bak%' and col.table_schema in('node','dbo')
                        and col.table_name not like '%backup%'
                        and col.table_name not like '%file_migration_list%'
                        order by col.table_schema,col.table_name,col.column_name
                """
            }
            host = [host for host in hosts if host.database == database][0].host
            pwd = [pwd for pwd in pwds if pwd.database == database][0].pwd
            user = [user for user in users if user.database == database][0].userName
            conn = psycopg2.connect(database=database, user=user,
                                    password=pwd, host=host, port=port)
            print("Opened database successfully")
            cur = conn.cursor()
            cur.execute(query["query"])
            data = cur.fetchall()

            queryedTable = []
            tableSet = list(set(map(lambda x: x[0], data)))
            tableCntDic = {}
            wb = openpyxl.Workbook()
            sheet = wb.active
            for index in range(len(cur.description)):
                # col_names.append()
                sheet.cell(row=1, column=index + 1,
                           value=cur.description[index].name)
            sheet.cell(row=1, column=len(cur.description) + 1, value="totalcnt")
            cur.close()
            for rowno, row in enumerate(tableSet, start=1):
                try:
                    cur = conn.cursor()
                    tableName = row
                    tableQuery = 'select count(1) from {}'.format(tableName)
                    if tableName in tableCntDic:
                        continue
                    cur.execute(tableQuery)
                    cnt: int = cur.fetchone()

                    tableCntDic[tableName] = cnt

                except Exception as e:
                    # 有些表没有id，会报错 catch之后直接跳过
                    print(e)
                    tableCntDic[tableName] = (-1,)
                    continue
                finally:
                    conn.rollback()
                    cur.close()

            for rowno, row in enumerate(data, start=1):

                tableName = row[0]

                cnt = tableCntDic[tableName]
                # 元组不能直接修改，只能创建新的
                completedRow = row + (cnt)
                for colno, cell_value in enumerate(completedRow, start=1):
                    # 处理NULL为NULL

                    if (cell_value == None):
                        sheet.cell(row=rowno + 1, column=colno, value='NULL')
                    else:  # 避免转换时候把空值转为字符串
                        if (isinstance(cell_value, str) and cell_value.lower() == "NULL".lower()):
                            sheet.cell(row=rowno + 1, column=colno,
                                       value='NULLStr')
                        else:  # 正常存储
                            if (type(cell_value) == bytes):
                                cell_value = "binary"
                            # 处理含有\ 特殊字符
                            # 处理特殊字符并去除添加的引号
                            if (isinstance(cell_value, str)):
                                cell_value = repr(cell_value)[1:-1]
                            # #bool to str
                            # if(isinstance(cell_value,bool)):
                            #     cell_value=str(cell_value)
                            sheet.cell(row=rowno + 1, column=colno,
                                       value=cell_value)
            sheet.row_dimensions[1].font = Font(bold=True)
            if query['name']:
                sheet.title = query['name']
            # 同名excel文件不让同时读取 即使在不同目录
            file_path = folder_path + "/" + query['name'] + '-pg.xlsx'
            print(file_path)
            wb.save(file_path)

            s3_key = "public/tmp/dbcompare/{}".format(database + '-' + query['name'] + '-pg.xlsx')
            upload_single_file_to_s3(
                s3_key=s3_key, file_path=file_path)
        except Exception as e:
            print(e)
        conn.close()


def test():
    with open("./hello_world/test_data.json", 'r', encoding="utf-8") as fq:
        # t=fq.readline()
        querys = json.load(fq)
        pathFolder = "C:/WorkSpace/db/data/pgsql"
        # export_to_excel(querys=querys, folder_path=pathFolder)
        export_to_db_schema(folder_path=pathFolder)


# test()


def lambda_handler(event, context):
    print(event)
    folder_path = "/tmp"
    export_to_excel(event, folder_path=folder_path)
    export_to_db_schema(folder_path=folder_path)
